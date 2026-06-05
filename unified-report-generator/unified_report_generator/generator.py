"""
unified_report_generator.generator — 统一报告生成引擎

包含：
- 台账加载、过滤、统计
- 交付文档提取与摘要
- Markdown / Word (.docx) 报告生成
- AI 增强报告（LLM 润色 + 统计图表）
"""

from collections import Counter
from datetime import date, datetime
from difflib import SequenceMatcher
from html import unescape
from pathlib import Path
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as xml_escape

import openpyxl

from .config import (
    TMP_REPORTS,
    REPORT_TYPES,
    WORK_SHEETS,
    DOCUMENT_EXTENSIONS,
    KEY_CONTENT_WORDS,
)
from ._llm import ask_llm


# ── 工具函数 ──────────────────────────────────────────────


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip().strip("'")
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass
    return None


def parse_iso_date(value, field_name):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception as exc:
        raise ValueError(f"{field_name} 必须是 YYYY-MM-DD 格式") from exc


def load_records(ledger_path: Path):
    if not ledger_path.exists() or not ledger_path.is_file():
        raise FileNotFoundError(f"Excel 台账不存在：{ledger_path}")

    wb = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
    records = []
    try:
        for sheet_name in WORK_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                continue
            header_names = [str(v).strip() if v is not None else "" for v in headers]
            last = max((i for i, header in enumerate(header_names) if header), default=-1)
            header_names = header_names[:last + 1]
            for row in rows:
                if not any(v is not None and str(v).strip() for v in row[:len(header_names)]):
                    continue
                record = {header_names[i]: row[i] if i < len(row) else None
                          for i in range(len(header_names)) if header_names[i]}
                record["数据来源"] = sheet_name
                record_date = parse_date(record.get("日期"))
                if not record_date:
                    continue
                record["_date"] = record_date
                records.append(record)
    finally:
        wb.close()
    return records


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def markdown_cell(value):
    return stringify(value).replace("\\", "\\\\").replace("|", "/").replace("\n", " ")


def normalize_doc_name(value):
    text = stringify(value)
    if not text:
        return ""
    return Path(text).name.strip()


# ── 交付文档索引与提取 ────────────────────────────────────


def build_document_index(document_root: Path):
    if not document_root.exists() or not document_root.is_dir():
        raise FileNotFoundError(f"交付文档搜索目录不存在：{document_root}")
    index = {}
    for path in document_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in DOCUMENT_EXTENSIONS:
            continue
        index.setdefault(path.name, []).append(path)
    return index


def find_delivery_document(document_name, document_index):
    name = normalize_doc_name(document_name)
    if not name:
        return None
    exact = document_index.get(name)
    if exact:
        return exact[0]
    stem = Path(name).stem
    candidates = []
    for indexed_name, paths in document_index.items():
        if stem and (stem in Path(indexed_name).stem or Path(indexed_name).stem in stem):
            candidates.extend(paths)
    if candidates:
        return candidates[0]

    date_match = re.search(r"\d{8}", stem)
    best_path = None
    best_score = 0
    for indexed_name, paths in document_index.items():
        indexed_stem = Path(indexed_name).stem
        if date_match and date_match.group(0) not in indexed_stem:
            continue
        score = SequenceMatcher(None, stem, indexed_stem).ratio()
        if score > best_score:
            best_score = score
            best_path = paths[0]
    return best_path if best_score >= 0.32 else None


def xml_iter_text(data):
    try:
        root = ET.fromstring(data)
    except Exception:
        text = data.decode("utf-8", errors="ignore")
        matches = re.findall(r"<(?:w|a|t):t[^>]*>(.*?)</(?:w|a|t):t>", text, flags=re.DOTALL)
        if not matches:
            matches = re.findall(r">([^<>]{2,})<", text)
        return [unescape(re.sub(r"<[^>]+>", "", item)).strip() for item in matches if item and item.strip()]
    return [text.strip() for text in root.itertext() if text and text.strip()]


def extract_docx_text(path: Path):
    chunks = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.startswith("word/") and name.endswith(".xml"):
                chunks.extend(xml_iter_text(archive.read(name)))
    return chunks


def extract_pptx_text(path: Path):
    chunks = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.startswith("ppt/") and name.endswith(".xml"):
                chunks.extend(xml_iter_text(archive.read(name)))
    return chunks


def extract_xlsx_text(path: Path):
    chunks = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [stringify(value) for value in row if stringify(value)]
                if values:
                    chunks.append(" | ".join(values))
    finally:
        workbook.close()
    return chunks


def extract_plain_text(path: Path):
    raw = path.read_bytes()
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return [line.strip() for line in raw.decode(encoding).splitlines() if line.strip()]
        except UnicodeDecodeError:
            continue
    return []


def extract_document_text(path: Path):
    ext = path.suffix.lower()
    if ext == ".docx":
        return extract_docx_text(path)
    if ext == ".pptx":
        return extract_pptx_text(path)
    if ext == ".xlsx":
        return extract_xlsx_text(path)
    if ext in {".txt", ".md", ".log"}:
        return extract_plain_text(path)
    return []


def summarize_document_chunks(chunks, limit=8):
    normalized = []
    seen = set()
    for chunk in chunks:
        text = re.sub(r"\s+", " ", stringify(chunk)).strip()
        if len(text) < 8 or text in seen:
            continue
        seen.add(text)
        normalized.append(text)
    key_lines = [line for line in normalized
                 if any(word.lower() in line.lower() for word in KEY_CONTENT_WORDS)]
    selected = (key_lines or normalized)[:limit]
    return [line[:220] for line in selected]


def collect_delivery_document_summaries(records, document_root: Path):
    index = build_document_index(document_root)
    summaries = []
    missing = []
    processed = set()
    for record in records:
        doc_name = normalize_doc_name(record.get("交付文档"))
        if not doc_name or doc_name in processed:
            continue
        processed.add(doc_name)
        path = find_delivery_document(doc_name, index)
        if not path:
            missing.append(doc_name)
            continue
        try:
            chunks = extract_document_text(path)
            summaries.append({
                "name": doc_name,
                "path": str(path),
                "summary": summarize_document_chunks(chunks),
            })
        except Exception as exc:
            summaries.append({
                "name": doc_name,
                "path": str(path),
                "summary": [f"文档解析失败：{exc}"],
            })
    return summaries, missing


# ── 统计与过滤 ────────────────────────────────────────────


def top_counts(records, field, limit=10):
    return Counter(stringify(r.get(field)) for r in records if stringify(r.get(field))).most_common(limit)


def filter_weekly_records(records, start_date, end_date):
    selected = [r for r in records if start_date <= r["_date"] <= end_date]
    reportable = [r for r in selected if stringify(r.get("纳入汇报")) != "否"]
    return selected, reportable


# ── Markdown 渲染辅助 ─────────────────────────────────────


def render_count_table(title, rows):
    lines = [f"### {title}", "", "| 项目 | 数量 |", "| --- | ---: |"]
    if rows:
        lines.extend(f"| {markdown_cell(name)} | {count} |" for name, count in rows)
    else:
        lines.append("| 无 | 0 |")
    lines.append("")
    return lines


def render_management_summary(report_config, start_date, end_date, reportable,
                               completed, continued, risk_records,
                               category_counts, system_counts, document_summaries):
    top_category = category_counts[0][0] if category_counts else "暂无"
    top_system = system_counts[0][0] if system_counts else "暂无"
    lines = [
        "## 管理摘要",
        "",
        (
            f"{report_config['title']}周期为 {start_date.isoformat()} 至 {end_date.isoformat()}。"
            f"本周期纳入汇报事项 {len(reportable)} 项，已完成 {len(completed)} 项，"
            f"延续事项 {len(continued)} 项，风险/影响事项 {len(risk_records)} 项。"
            f"主要工作类别为 {top_category}，主要涉及系统为 {top_system}。"
            f"本次关联并整理交付文档 {len(document_summaries)} 份。"
        ),
        "",
    ]
    return lines


def render_record_table(title, records, limit=20):
    lines = [f"### {title}", ""]
    if not records:
        lines.extend(["无。", ""])
        return lines
    lines.extend([
        "| 日期 | 来源 | 类别 | 系统 | 事件级别 | 业务影响 | 结果 | 工作内容 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ])
    for r in records[:limit]:
        content = markdown_cell(r.get("工作内容"))[:120]
        lines.append(
            "| {date} | {source} | {category} | {system} | {level} | {impact} | {result} | {content} |".format(
                date=r["_date"].isoformat(),
                source=markdown_cell(r.get("数据来源")),
                category=markdown_cell(r.get("类别")),
                system=markdown_cell(r.get("系统名称")),
                level=markdown_cell(r.get("事件级别")) or "-",
                impact=markdown_cell(r.get("业务影响")) or "-",
                result=markdown_cell(r.get("结果")) or "-",
                content=content,
            )
        )
    lines.append("")
    return lines


def render_delivery_document_list(title, summaries, missing):
    lines = [f"### {title}", ""]
    if not summaries and not missing:
        lines.extend(["无交付文档记录。", ""])
        return lines
    lines.extend(["| 文档名 | 状态 | 路径 |", "| --- | --- | --- |"])
    for item in summaries:
        lines.append(f"| {markdown_cell(item['name'])} | 已找到 | `{markdown_cell(item['path'])}` |")
    for name in missing:
        lines.append(f"| {markdown_cell(name)} | 未找到 | - |")
    lines.append("")
    return lines


def render_delivery_document_summaries(title, summaries, missing):
    lines = [f"### {title}", ""]
    if not summaries and not missing:
        lines.extend(["本周期台账记录未填写交付文档。", ""])
        return lines
    for item in summaries:
        lines.extend([
            f"#### {markdown_cell(item['name'])}",
            "",
            f"- 文件路径：`{markdown_cell(item['path'])}`",
            "- 重点内容：",
        ])
        if item["summary"]:
            lines.extend(f"  - {markdown_cell(line)}" for line in item["summary"])
        else:
            lines.append("  - 未提取到有效文本。")
        lines.append("")
    if missing:
        lines.extend(["#### 未找到的交付文档", ""])
        lines.extend(f"- {markdown_cell(name)}" for name in missing)
        lines.append("")
    return lines


# ── Word / OpenXML 渲染 ──────────────────────────────────


def word_text(text):
    return xml_escape(stringify(text), {'"': '"'})


def word_color_hex(r, g, b):
    return f"{r:02X}{g:02X}{b:02X}"


# 品牌色
BRAND_TEAL = "0D9488"
BRAND_DARK = "0F766E"
BRAND_LIGHT = "CCFBF1"
HEADER_BG = "0F766E"
ROW_ALT_BG = "F0FDFA"
ACCENT_AMBER = "D97706"
ACCENT_RED = "DC2626"
ACCENT_GREEN = "15803D"


def word_run(text, bold=False, italic=False, size=22, color=None, font=None):
    props = []
    if bold:
        props.append('<w:b/>')
    if italic:
        props.append('<w:i/>')
    props.append(f'<w:sz w:val="{size}"/>')
    props.append(f'<w:szCs w:val="{size}"/>')
    if color:
        props.append(f'<w:color w:val="{color}"/>')
    if font:
        props.append(f'<w:rFonts w:ascii="{font}" w:hAnsi="{font}" w:eastAsia="{font}"/>')
    else:
        props.append('<w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>')
    run_props = '<w:rPr>' + ''.join(props) + '</w:rPr>'
    return f'<w:r>{run_props}<w:t xml:space="preserve">{word_text(text)}</w:t></w:r>'


def word_paragraph(text="", bold=False, italic=False, size=22, color=None,
                   alignment=None, spacing_before=0, spacing_after=120, font=None):
    pPr_parts = []
    if alignment == 'center':
        pPr_parts.append('<w:jc w:val="center"/>')
    elif alignment == 'right':
        pPr_parts.append('<w:jc w:val="right"/>')
    elif alignment == 'left':
        pPr_parts.append('<w:jc w:val="left"/>')
    if spacing_before is not None or spacing_after is not None:
        sp = '<w:spacing'
        if spacing_before is not None:
            sp += f' w:before="{spacing_before}"'
        if spacing_after is not None:
            sp += f' w:after="{spacing_after}"'
        sp += '/>'
        pPr_parts.append(sp)
    pPr = f'<w:pPr>{"".join(pPr_parts)}</w:pPr>' if pPr_parts else ''
    run = word_run(text, bold=bold, italic=italic, size=size, color=color, font=font)
    return f'<w:p>{pPr}{run}</w:p>'


def word_heading(text, level=1):
    sizes = {1: 44, 2: 32, 3: 26, 4: 22}
    size = sizes.get(level, 24)
    colors = {1: BRAND_DARK, 2: BRAND_TEAL, 3: '333333', 4: '555555'}
    color = colors.get(level, '333333')
    before = {1: 360, 2: 240, 3: 200, 4: 160}.get(level, 120)
    after = {1: 200, 2: 160, 3: 120, 4: 80}.get(level, 80)
    return word_paragraph(text, bold=True, size=size, color=color,
                          spacing_before=before, spacing_after=after)


def word_bullet(text, indent=720, hanging=360):
    return (
        f'<w:p><w:pPr><w:ind w:left="{indent}" w:hanging="{hanging}"/></w:pPr>'
        f'<w:r><w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>'
        f'<w:sz w:val="22"/></w:rPr>'
        f'<w:t xml:space="preserve">• {word_text(text)}</w:t></w:r></w:p>'
    )


def word_page_break():
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def word_horizontal_rule():
    return '<w:p><w:pPr><w:pBdr><w:bottom w:val="single" w:sz="6" w:space="1" w:color="CCCCCC"/></w:pBdr></w:pPr></w:p>'


def word_empty_line():
    return '<w:p><w:pPr><w:spacing w:after="60"/></w:pPr></w:p>'


def word_shading(color_hex):
    return f'<w:shd w:val="clear" w:color="auto" w:fill="{color_hex}"/>'


def word_table_cell(value, bold=False, width=2400, shading=None, alignment=None, size=20, color=None):
    tc_pr_parts = [f'<w:tcW w:w="{width}" w:type="dxa"/>']
    if shading:
        tc_pr_parts.append(word_shading(shading))
    tc_pr = '<w:tcPr>' + ''.join(tc_pr_parts) + '</w:tcPr>'
    p_align = ''
    if alignment == 'center':
        p_align = '<w:jc w:val="center"/>'
    elif alignment == 'right':
        p_align = '<w:jc w:val="right"/>'
    run = word_run(value, bold=bold, size=size, color=color)
    p = f'<w:p><w:pPr>{p_align}<w:spacing w:before="40" w:after="40"/></w:pPr>{run}</w:p>'
    return f'<w:tc>{tc_pr}{p}</w:tc>'


def word_table(headers, rows):
    col_count = len(headers)
    col_width = 9000 // col_count
    table = [
        '<w:tbl>',
        '<w:tblPr>',
        '<w:tblStyle w:val="TableGrid"/>',
        '<w:tblW w:w="9000" w:type="dxa"/>',
        '<w:jc w:val="center"/>',
        '<w:tblBorders>',
        '<w:top w:val="single" w:sz="8" w:space="0" w:color="0D9488"/>',
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>',
        '<w:bottom w:val="single" w:sz="8" w:space="0" w:color="0D9488"/>',
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="CCCCCC"/>',
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="DDDDDD"/>',
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="DDDDDD"/>',
        '</w:tblBorders>',
        '</w:tblPr>',
    ]
    # 表头行
    table.append('<w:tr>')
    for i, header in enumerate(headers):
        width = col_width if i < col_count - 1 else 9000 - col_width * (col_count - 1)
        table.append(word_table_cell(header, bold=True, width=width,
                                     shading=HEADER_BG, alignment='center', size=20, color='FFFFFF'))
    table.append('</w:tr>')
    # 数据行
    for row_idx, row in enumerate(rows):
        shading = ROW_ALT_BG if row_idx % 2 == 1 else None
        table.append('<w:tr>')
        for i, value in enumerate(row):
            width = col_width if i < col_count - 1 else 9000 - col_width * (col_count - 1)
            is_numeric = isinstance(value, (int, float)) or (isinstance(value, str) and value.replace(',', '').isdigit())
            align = 'right' if is_numeric and i > 0 else 'left'
            table.append(word_table_cell(str(value), width=width, shading=shading, alignment=align, size=20))
        table.append('</w:tr>')
    table.append('</w:tbl>')
    return ''.join(table)


def word_info_row(label, value, color=None):
    return (
        f'<w:p><w:pPr><w:spacing w:before="40" w:after="40"/></w:pPr>'
        f'{word_run(label, bold=True, size=20, color="666666")}'
        f'{word_run("  " + value, size=20, color=color or "333333")}'
        f'</w:p>'
    )


def word_badge(text, bg_color="0D9488", text_color="FFFFFF"):
    return (
        f'<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
        f'<w:r><w:rPr><w:rFonts w:ascii="Microsoft YaHei" w:hAnsi="Microsoft YaHei" w:eastAsia="Microsoft YaHei"/>'
        f'<w:b/><w:sz w:val="18"/><w:color w:val="{text_color}"/>'
        f'<w:shd w:val="clear" w:color="auto" w:fill="{bg_color}"/></w:rPr>'
        f'<w:t xml:space="preserve"> {word_text(text)} </w:t></w:r></w:p>'
    )


# ── Word 报告生成 ────────────────────────────────────────


def create_docx_report(output_path: Path, context: dict):
    body = []
    report_config = context["report_config"]
    start_date = context["start_date"]
    end_date = context["end_date"]
    today_str = datetime.now().strftime("%Y年%m月%d日")

    # ========== 封 面 ==========
    for _ in range(6):
        body.append(word_empty_line())
    body.append(word_paragraph("服务交付", bold=True, size=56, color="0D9488",
                               alignment='center', spacing_before=0, spacing_after=0))
    body.append(word_empty_line())
    body.append(word_paragraph(f"{report_config['title']}", bold=True, size=48,
                               color="111C34", alignment='center', spacing_before=0, spacing_after=120))
    body.append(word_paragraph("━" * 30, size=18, color="0D9488",
                               alignment='center', spacing_before=200, spacing_after=200))
    body.append(word_paragraph(f"{start_date.isoformat()} 至 {end_date.isoformat()}",
                               size=28, color="555555", alignment='center',
                               spacing_before=60, spacing_after=240))
    info_items = [
        ("客  户", "吉林银行"),
        ("报告周期", f"{start_date.isoformat()} — {end_date.isoformat()}"),
        ("生成日期", today_str),
        ("数据来源", str(context["ledger"].name)),
    ]
    for label, value in info_items:
        body.append(word_info_row(label, value, color="333333") if label else word_empty_line())
    body.append(word_empty_line())
    body.append(word_empty_line())
    body.append(word_paragraph("Oracle 系统集成中心 · 服务交付报告", size=18,
                               color="999999", alignment='center', spacing_before=240, spacing_after=60))
    body.append(word_paragraph("本报告为内部资料，仅供参考", size=16,
                               color="AAAAAA", alignment='center', spacing_before=0, spacing_after=400))
    body.append(word_page_break())

    # ========== 目 录 ==========
    body.append(word_heading("目  录", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    toc_items = [
        ("一、服务概览", 2),
        ("二、管理摘要", 2),
        ("三、分类统计", 2),
        ("   3.1 按类别统计", 3),
        ("   3.2 按系统统计", 3),
        ("   3.3 按人员分布", 3),
        ("四、重点事项", 2),
        ("五、风险与问题", 2),
        ("六、交付文档重点内容", 2),
        ("七、后续计划", 2),
        ("八、交付文档清单", 2),
    ]
    for item_text, level in toc_items:
        sz = {2: 24, 3: 22}.get(level, 22)
        indent = {2: 0, 3: 360}.get(level, 0)
        color = {2: "333333", 3: "666666"}.get(level, "666666")
        ppr = f'<w:pPr><w:ind w:left="{indent}"/><w:spacing w:before="80" w:after="80"/></w:pPr>'
        run = word_run(item_text, size=sz, color=color)
        body.append(f'<w:p>{ppr}{run}</w:p>')
    body.append(word_empty_line())
    body.append(word_horizontal_rule())
    body.append(word_page_break())

    # ========== 正 文 ==========
    # 一、服务概览
    body.append(word_heading("一、服务概览", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    overview_data = [
        ["周期内记录数", str(context["selected_count"])],
        ["纳入汇报记录数", str(len(context["reportable"]))],
        ["已完成事项", str(len(context["completed"]))],
        ["延续事项", str(len(context["continued"]))],
        ["风险/影响事项", str(len(context["risk_records"]))],
        ["交付文档数量", str(len(context["document_summaries"]))],
    ]
    body.append(word_table(["指标", "数量"], overview_data))
    body.append(word_empty_line())
    # 二、管理摘要
    body.append(word_heading("二、管理摘要", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    top_category = context["category_counts"][0][0] if context["category_counts"] else "暂无"
    top_system = context["system_counts"][0][0] if context["system_counts"] else "暂无"
    summary_text = (
        f"本{report_config['title']}周期为 {start_date.isoformat()} 至 {end_date.isoformat()}。"
        f"本周期纳入汇报事项 {len(context['reportable'])} 项，已完成 {len(context['completed'])} 项，"
        f"延续事项 {len(context['continued'])} 项，风险/影响事项 {len(context['risk_records'])} 项。"
        f"主要工作类别为「{top_category}」，主要涉及系统为「{top_system}」。"
        f"本次关联并整理交付文档 {len(context['document_summaries'])} 份。"
    )
    body.append(word_paragraph(summary_text, size=22, color="333333",
                               spacing_before=60, spacing_after=200))
    # 三、分类统计
    body.append(word_heading("三、分类统计", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    body.append(word_heading("3.1 按类别统计", 3))
    body.append(word_table(["类别", "数量"], context["category_counts"] or [["无", "0"]]))
    body.append(word_empty_line())
    body.append(word_heading("3.2 按系统统计", 3))
    body.append(word_table(["系统", "数量"], context["system_counts"] or [["无", "0"]]))
    body.append(word_empty_line())
    body.append(word_heading("3.3 按人员分布", 3))
    body.append(word_table(["人员", "数量"], context["person_counts"] or [["无", "0"]]))
    body.append(word_empty_line())
    # 四、重点事项
    body.append(word_heading(f"四、{report_config['focus_title']}", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    focus_rows = [
        [
            r["_date"].isoformat(),
            stringify(r.get("类别")),
            stringify(r.get("系统名称")),
            stringify(r.get("结果")) or "-",
            stringify(r.get("工作内容"))[:180],
        ]
        for r in context["reportable"][:30]
    ]
    body.append(word_table(["日期", "类别", "系统", "结果", "工作内容"],
                           focus_rows or [["-", "-", "-", "-", "无"]]))
    body.append(word_empty_line())
    # 五、风险与问题
    body.append(word_heading("五、风险与问题", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    if context["risk_records"]:
        body.append(word_paragraph(f"共 {len(context['risk_records'])} 项风险或影响事项：",
                                   size=22, color="DC2626", spacing_before=60, spacing_after=120))
        for record in context["risk_records"][:20]:
            body.append(word_bullet(
                f"{record['_date'].isoformat()} {stringify(record.get('系统名称'))}："
                f"{stringify(record.get('工作内容'))[:180]}",
                indent=480, hanging=240))
    else:
        body.append(word_paragraph("本周期无风险或影响事项。", size=22,
                                   color="15803D", spacing_before=60))
    body.append(word_empty_line())
    # 六、交付文档重点内容
    body.append(word_heading("六、交付文档重点内容", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    if context["document_summaries"]:
        for item in context["document_summaries"]:
            body.append(word_heading(item["name"], 3))
            body.append(word_paragraph(f"文件路径：{item['path']}", size=18,
                                       color="888888", spacing_before=40, spacing_after=60))
            for line in item["summary"][:8]:
                body.append(word_bullet(line, indent=480, hanging=240))
            body.append(word_empty_line())
    else:
        body.append(word_paragraph("本周期台账记录未填写或未匹配到交付文档。",
                                   size=22, color="888888", spacing_before=60))
    body.append(word_empty_line())
    # 七、后续计划
    body.append(word_heading(f"七、{report_config['plan_title']}", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    for item in report_config["plan_items"]:
        body.append(word_bullet(item, indent=480, hanging=240))
    body.append(word_empty_line())
    # 八、交付文档清单
    body.append(word_heading("八、交付文档清单", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    list_rows = [[item["name"], "已找到", item["path"]] for item in context["document_summaries"]]
    list_rows.extend([[name, "未找到", "-"] for name in context["missing_documents"]])
    body.append(word_table(["文档名", "状态", "路径"], list_rows or [["无", "-", "-"]]))
    body.append(word_empty_line())
    # 页脚
    body.append(word_empty_line())
    body.append(word_horizontal_rule())
    body.append(word_paragraph(f"生成日期：{today_str} | 数据来源：{context['ledger'].name}",
                               size=16, color="AAAAAA", alignment='center',
                               spacing_before=120, spacing_after=0))

    # ========== 文档主 XML 结构 ==========
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body>'
        + "".join(body)
        + '<w:sectPr>'
        '<w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1440" w:right="1200" w:bottom="1440" w:left="1200"/>'
        '<w:cols w:space="720"/>'
        '</w:sectPr>'
        '</w:body></w:document>'
    )

    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '</Relationships>'
    )
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('[Content_Types].xml', content_types)
        archive.writestr('_rels/.rels', rels)
        archive.writestr('word/document.xml', document_xml)


# ── AI 增强报告 ──────────────────────────────────────────


def build_report_prompt(context: dict, lines: list[str]) -> str:
    """构建发送给大模型的 Prompt，包含结构化统计数据和原始 Markdown。"""
    report_config = context["report_config"]
    start_date = context["start_date"]
    end_date = context["end_date"]

    stats_summary = {
        "report_type": report_config["title"],
        "period": f"{start_date.isoformat()} ~ {end_date.isoformat()}",
        "total_records": context["selected_count"],
        "reportable_records": len(context["reportable"]),
        "completed": len(context["completed"]),
        "continued": len(context["continued"]),
        "risk_count": len(context["risk_records"]),
        "category_counts": [(k, v) for k, v in context["category_counts"]],
        "system_counts": [(k, v) for k, v in context["system_counts"]],
        "person_counts": [(k, v) for k, v in context["person_counts"]],
        "source_counts": [(k, v) for k, v in context["source_counts"]],
        "result_counts": [(k, v) for k, v in context["result_counts"]],
        "delivery_documents": len(context["document_summaries"]),
    }

    key_records = []
    for r in context["reportable"][:20]:
        key_records.append({
            "date": r["_date"].isoformat(),
            "source": stringify(r.get("数据来源")),
            "category": stringify(r.get("类别")),
            "system": stringify(r.get("系统名称")),
            "level": stringify(r.get("事件级别")),
            "impact": stringify(r.get("业务影响")),
            "result": stringify(r.get("结果")),
            "content": stringify(r.get("工作内容"))[:200],
        })

    risk_summary = []
    for r in context["risk_records"][:15]:
        risk_summary.append({
            "date": r["_date"].isoformat(),
            "system": stringify(r.get("系统名称")),
            "level": stringify(r.get("事件级别")),
            "impact": stringify(r.get("业务影响")),
            "content": stringify(r.get("工作内容"))[:200],
        })

    doc_summaries_text = ""
    for item in context["document_summaries"][:10]:
        doc_summaries_text += f"\n#### {item['name']}\n"
        for line in item["summary"][:10]:
            doc_summaries_text += f"- {line}\n"

    original_md = "\n".join(lines[:200])

    prompt = f"""你是一位资深的金融行业 Oracle 数据库服务交付专家，负责撰写专业的企业级服务交付报告。

## 报告数据上下文

```json
{json.dumps(stats_summary, ensure_ascii=False, indent=2)}
```

## 详细工作记录

```json
{json.dumps(key_records, ensure_ascii=False, indent=2)}
```

## 风险事项摘要

```json
{json.dumps(risk_summary, ensure_ascii=False, indent=2) if risk_summary else "无"}
```

## 交付文档摘要

{doc_summaries_text if doc_summaries_text else "无交付文档"}

## 原始 Markdown

以下是当前程序自动生成的原始报告 Markdown，请在此基础之上进行优化润色：

```
{original_md}
```

## 指令要求

### 总体要求
输出**增强版 Markdown**，保持原有的 12 节章节结构完整，每节必须包含：

1. **数据表格** — 用 Markdown 表格呈现核心数据
2. **文字洞察** — 对数据进行分析解读，不少于 100 字
3. **趋势/结论** — 给出明确的结论或改进建议

### 各节具体要求

#### 1. 管理摘要（重写为专业分析报告风格）
- 用金融行业专业语言撰写，不少于 300 字
- 包含：本期工作概况、重点成果、风险警示、改进方向
- 格式：段落 + 关键数据引用

#### 2. 概览（保留原有数据 → 添加解读表格）
- 保留所有的 KPI 数据行
- 在数据下方添加一个「关键指标解读表」：

| 指标 | 数值 | 解读 |
| 周期内记录数 | XX | 工作负载情况说明 |
| 已完成事项 | XX | 完成率 XX% |

#### 3. 按类别统计（表格 + 饼图标记 + 分析）
- 保留原有统计表格
- 插入饼图标记：`<!--CHART:pie:工作类别分布-->`
- 添加文字分析：哪些类别占比最高？为什么？

#### 4. 按系统统计（表格 + 柱状图标记 + 分析）
- 保留原有统计表格
- 插入柱状图标记：`<!--CHART:bar:系统工作量分布-->`
- 添加文字分析：哪些系统是重点？趋势如何？

#### 5. 按人员分布（表格 + 柱状图标记 + 分析）
- 保留原有统计表格
- 插入柱状图标记：`<!--CHART:bar:人员工作分布-->`
- 添加文字分析：人员工作负载是否均衡

#### 6. 重点事项（增强表格 + 分析）
- 每项记录展开为完整描述
- 添加「重点关注」标记（高/中/普通）

#### 7. 风险与问题（表格 + 风险分析）
- 列出风险事项后，添加「风险分析表」：

| 风险等级 | 涉及系统 | 影响范围 | 建议措施 |
|----------|----------|----------|----------|

- 给出具体的改进建议和责任人建议

#### 8. 交付文档（文档摘要 + 类型统计）
- 列出交付文档后，添加文档类型统计

#### 9. 后续计划
- 保留原有计划项
- 添加基于风险分析的额外建议

### 图表占位标记
在需要插入图表的位置使用以下标记（程序会自动替换为真实图表）：
- `<!--CHART:pie:工作类别分布-->`
- `<!--CHART:bar:系统工作量分布-->`
- `<!--CHART:bar:人员工作分布-->`
- `<!--CHART:pie:结果状态分布-->`

### 格式约束
- 使用标准 Markdown 语法
- 表格使用 `|` 分隔
- 标题使用 `##` 和 `###`
- 所有数字用阿拉伯数字
- 不包含额外的说明文字（如"以下是优化后的报告"）
- 直接输出增强后的 Markdown 内容"""
    return prompt


# ── 图表生成 ─────────────────────────────────────────────


def generate_report_charts(context: dict, output_dir: Path) -> list[str]:
    """基于统计数据生成图表 PNG，返回图表文件路径列表。"""
    chart_paths = []

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
    except ImportError:
        return chart_paths

    cn_font = None
    for candidate in ["PingFang SC", "PingFang HK", "Microsoft YaHei",
                       "WenQuanYi Micro Hei", "Noto Sans CJK SC", "SimHei"]:
        try:
            fm.findfont(candidate, fallback_to_default=False)
            cn_font = candidate
            break
        except Exception:
            continue

    plt.rcParams["font.family"] = cn_font or "sans-serif"
    plt.rcParams["axes.unicode_minus"] = False

    DEEP_BLUE = "#1B3A5C"
    GOLD = "#C8A96E"
    TEAL = "#0D9488"
    RED = "#DC2626"
    GREEN = "#15803D"
    GRAY = "#94A3B8"
    COLORS_PIE = [DEEP_BLUE, GOLD, TEAL, GREEN, RED, GRAY, "#8B5CF6", "#F59E0B"]
    COLORS_BAR = [DEEP_BLUE, GOLD, TEAL, GREEN, "#8B5CF6"]

    def _pie(data, title, filename):
        if not data:
            return
        labels = [item[0] if len(item) > 4 else item[0][:8] for item in data]
        values = [item[1] for item in data]
        fig, ax = plt.subplots(figsize=(6, 4.5))
        colors = COLORS_PIE[:len(data)]
        wedges, texts, autotexts = ax.pie(
            values, labels=None, autopct="%1.1f%%",
            colors=colors, startangle=90,
            textprops={"fontsize": 9},
            pctdistance=0.75,
        )
        ax.legend(wedges, labels, title=title, loc="center left",
                  bbox_to_anchor=(1, 0, 0.5, 1), fontsize=8)
        ax.set_title(title, fontsize=12, fontweight="bold", color=DEEP_BLUE, pad=12)
        plt.tight_layout()
        path = output_dir / filename
        fig.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        chart_paths.append(str(path))

    def _bar(data, title, filename):
        if not data:
            return
        labels = [item[0][:6] for item in data]
        values = [item[1] for item in data]
        fig, ax = plt.subplots(figsize=(6, 4.5))
        bars = ax.bar(labels, values, color=COLORS_BAR[:len(data)],
                      edgecolor="white", linewidth=0.5)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3, str(val),
                    ha="center", va="bottom", fontsize=9, fontweight="bold", color=DEEP_BLUE)
        ax.set_title(title, fontsize=12, fontweight="bold", color=DEEP_BLUE, pad=12)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(axis="x", labelsize=8)
        ax.tick_params(axis="y", labelsize=8)
        plt.tight_layout()
        path = output_dir / filename
        fig.savefig(path, dpi=200, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        chart_paths.append(str(path))

    _pie(context["category_counts"], "工作类别分布", "chart_category_pie.png")
    _bar(context["system_counts"], "系统工作量分布", "chart_system_bar.png")
    _bar(context["person_counts"], "人员工作分布", "chart_person_bar.png")
    result_data = [
        ["完成", len(context["completed"])],
        ["延续", len(context["continued"])],
        ["风险", len(context["risk_records"])],
    ]
    _pie(result_data, "结果状态分布", "chart_result_pie.png")

    return chart_paths


def _image_to_word_xml(image_path: str, max_width_emu: int = 4500000) -> tuple:
    """将 PNG 图片转换为 OpenXML wp:inline 嵌入元素。"""
    from base64 import b64encode
    import struct
    import zlib

    path = Path(image_path)
    if not path.exists():
        return "", 0, ""

    with open(path, "rb") as f:
        f.read(8)
        ihdr_data = f.read(25)
        width_px, height_px = struct.unpack(">II", ihdr_data[8:16])

    with open(path, "rb") as f:
        img_data = f.read()
    img_b64 = b64encode(img_data).decode("ascii")

    dpi = 200
    width_emu = min(int(width_px / dpi * 914400), max_width_emu)
    height_emu = int(height_px / dpi * 914400)

    img_id = abs(hash(image_path)) % 100000
    rel_id = img_id

    return (
        f'<w:p><w:r><w:drawing>'
        f'<wp:inline distT="0" distB="0" distL="0" distR="0">'
        f'<wp:extent cx="{width_emu}" cy="{height_emu}"/>'
        f'<wp:effectExtent l="0" t="0" r="0" b="0"/>'
        f'<wp:docPr id="{img_id}" name="Chart{img_id}"/>'
        f'<wp:cNvGraphicFramePr><a:graphicFrameLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/></wp:cNvGraphicFramePr>'
        f'<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f'<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:nvPicPr><pic:cNvPr id="{img_id}" name="Chart{img_id}"/><pic:cNvPicPr/></pic:nvPicPr>'
        f'<pic:blipFill><a:blip r:embed="rId{rel_id}" cstate="print"><a:extLst><a:ext uri="{{28D789D4-4E23-4B1D-B63F-2F8AD8C7B0B5}}"><a14:useLocalDpi xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" val="0"/></a:ext></a:extLst></a:blip>'
        f'<a:srcRect/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
        f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{width_emu}" cy="{height_emu}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></a:spPr>'
        f'</pic:pic>'
        f'</a:graphicData>'
        f'</a:graphic>'
        f'</wp:inline>'
        f'</w:drawing></w:r></w:p>'
    ), rel_id, img_b64


def _replace_chart_placeholders(markdown_text: str, chart_paths: list[str]) -> tuple:
    result_lines = []
    rel_entries = []
    image_blobs = {}
    chart_idx = 0

    for line in markdown_text.splitlines():
        match = re.match(r"<!--CHART:(pie|bar):(.+?)-->", line.strip())
        if match and chart_idx < len(chart_paths):
            img_path = chart_paths[chart_idx]
            chart_idx += 1
            try:
                xml_part, rel_id, img_b64 = _image_to_word_xml(img_path)
                result_lines.append(f"__IMAGE__{rel_id}__")
                rel_entries.append((rel_id, "image/png", f"media/chart_{rel_id}.png"))
                image_blobs[f"word/media/chart_{rel_id}.png"] = img_b64
            except Exception:
                result_lines.append(line)
        else:
            result_lines.append(line)

    return "\n".join(result_lines), rel_entries, image_blobs


def create_ai_docx_report(output_path: Path, enhanced_markdown: str,
                           context: dict, chart_paths: list[str]):
    """基于 LLM 增强的 Markdown 生成金融版 Word 文档，含内嵌图表。"""
    body = []
    report_config = context["report_config"]
    start_date = context["start_date"]
    end_date = context["end_date"]
    today_str = datetime.now().strftime("%Y年%m月%d日")

    # ========== 封 面 ==========
    for _ in range(6):
        body.append(word_empty_line())
    body.append(word_paragraph("服务交付", bold=True, size=56, color="0D9488",
                               alignment='center', spacing_before=0, spacing_after=0))
    body.append(word_empty_line())
    body.append(word_paragraph(f"{report_config['title']}", bold=True, size=48,
                               color="111C34", alignment='center', spacing_before=0, spacing_after=120))
    body.append(word_paragraph("━" * 30, size=18, color="0D9488",
                               alignment='center', spacing_before=200, spacing_after=200))
    body.append(word_paragraph(f"{start_date.isoformat()} 至 {end_date.isoformat()}",
                               size=28, color="555555", alignment='center',
                               spacing_before=60, spacing_after=240))
    info_items = [
        ("客  户", "吉林银行"),
        ("报告周期", f"{start_date.isoformat()} — {end_date.isoformat()}"),
        ("生成日期", today_str),
        ("数据来源", str(context["ledger"].name)),
    ]
    for label, value in info_items:
        body.append(word_info_row(label, value, color="333333") if label else word_empty_line())
    body.append(word_empty_line())
    body.append(word_empty_line())
    body.append(word_paragraph("Oracle 系统集成中心 · 服务交付报告", size=18,
                               color="999999", alignment='center', spacing_before=240, spacing_after=60))
    body.append(word_paragraph("本报告为内部资料 · AI 辅助生成", size=16,
                               color="AAAAAA", alignment='center', spacing_before=0, spacing_after=400))
    body.append(word_page_break())

    # ========== 目 录 ==========
    body.append(word_heading("目  录", 1))
    body.append(word_horizontal_rule())
    body.append(word_empty_line())
    toc_items = [
        ("一、服务概览", 2), ("二、管理摘要（AI 增强）", 2),
        ("三、分类统计", 2), ("   3.1 按类别统计", 3), ("   3.2 按系统统计", 3), ("   3.3 按人员分布", 3),
        ("四、重点事项", 2), ("五、风险与问题", 2),
        ("六、交付文档重点内容", 2), ("七、后续计划", 2), ("八、交付文档清单", 2),
    ]
    for item_text, level in toc_items:
        sz = {2: 24, 3: 22}.get(level, 22)
        indent = {2: 0, 3: 360}.get(level, 0)
        color = {2: "333333", 3: "666666"}.get(level, "666666")
        ppr = f'<w:pPr><w:ind w:left="{indent}"/><w:spacing w:before="80" w:after="80"/></w:pPr>'
        run = word_run(item_text, size=sz, color=color)
        body.append(f'<w:p>{ppr}{run}</w:p>')
    body.append(word_empty_line())
    body.append(word_horizontal_rule())
    body.append(word_page_break())

    # ========== 正 文 ==========
    md_body, rel_entries, image_blobs = _replace_chart_placeholders(enhanced_markdown, chart_paths)

    for raw_line in md_body.splitlines():
        line = raw_line.strip()
        if not line:
            body.append(word_empty_line())
            continue
        if line.startswith("__IMAGE__") and line.endswith("__"):
            continue
        if re.match(r"<!--CHART:", line):
            body.append(word_paragraph(f"[图表：{line}]", italic=True, size=18, color="888888"))
            continue
        if line.startswith("### "):
            body.append(word_heading(line[4:].strip(), 3))
        elif line.startswith("## "):
            body.append(word_heading(line[3:].strip(), 2))
        elif line.startswith("# "):
            body.append(word_heading(line[2:].strip(), 1))
        elif line.startswith(("- ", "* ")):
            body.append(word_bullet(line[2:].strip()))
        elif re.match(r"^\d+[.、]\s+", line):
            body.append(word_bullet(line))
        elif line.startswith("| "):
            body.append(word_paragraph(line, size=20, color="333333"))
        else:
            body.append(word_paragraph(line))

    if chart_paths:
        body.append(word_heading("附：统计图表", 2))
        body.append(word_horizontal_rule())
        body.append(word_empty_line())
        for i, img_path in enumerate(chart_paths):
            try:
                xml_part, rel_id, img_b64 = _image_to_word_xml(img_path)
                body.append(xml_part)
                body.append(word_empty_line())
            except Exception:
                body.append(word_paragraph(f"[图表 {i+1}]", italic=True, size=18, color="888888"))

    body.append(word_empty_line())
    body.append(word_horizontal_rule())
    body.append(word_paragraph(
        f"生成日期：{today_str} | 数据来源：{context['ledger'].name} | AI 辅助生成",
        size=16, color="AAAAAA", alignment='center', spacing_before=120, spacing_after=0))

    sect_pr = ('<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
               '<w:pgMar w:top="1440" w:right="1200" w:bottom="1440" w:left="1200"/>'
               '<w:cols w:space="720"/></w:sectPr>')
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<w:body>'
        + "".join(body)
        + sect_pr
        + '</w:body></w:document>'
    )

    content_types_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
    ]
    for rel_id, ct, target in rel_entries:
        ext = target.split(".")[-1]
        content_types_parts.append(f'<Default Extension="{ext}" ContentType="{ct}"/>')
    content_types_parts.extend([
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>',
        '</Types>',
    ])
    content_types = "".join(content_types_parts)

    rels_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>',
    ]
    for rel_id, ct, target in rel_entries:
        rels_parts.append(
            f'<Relationship Id="rId{rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"'
            f' Target="word/{target}"/>')
    rels_parts.append('</Relationships>')
    rels = "".join(rels_parts)

    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('[Content_Types].xml', content_types)
        archive.writestr('_rels/.rels', rels)
        archive.writestr('word/document.xml', document_xml)
        for blob_path, b64_data in image_blobs.items():
            import base64
            archive.writestr(blob_path, base64.b64decode(b64_data))


# ── 报告主入口 ───────────────────────────────────────────


def generate_report_with_ai(
    report_type, ledger_path, start_date_text, end_date_text,
    document_root_text=None, url="", model="", api_key=""
):
    """生成 AI 增强报告：统计 + LLM 润色 + 图表 + 金融版 Word。"""
    report_config = REPORT_TYPES.get(report_type)
    if not report_config:
        raise ValueError("报告类型必须是 weekly、monthly、quarterly 或 annual")

    ledger = Path(ledger_path).expanduser().resolve()
    if document_root_text:
        document_root_text = document_root_text.strip().strip("'\"")
    document_root = Path(document_root_text).expanduser().resolve() if document_root_text else ledger.parent
    start_date = parse_iso_date(start_date_text, "开始日期")
    end_date = parse_iso_date(end_date_text, "结束日期")
    if end_date < start_date:
        raise ValueError("结束日期不能早于开始日期")

    records = load_records(ledger)
    selected, reportable = filter_weekly_records(records, start_date, end_date)
    completed = [r for r in reportable if stringify(r.get("结果")) == "完成"]
    continued = [r for r in reportable if stringify(r.get("结果")) == "延续"]
    risk_records = [r for r in reportable
                    if stringify(r.get("业务影响")) in {"已影响", "有风险"}
                    or stringify(r.get("事件级别")) in {"高", "中"}]

    category_counts = top_counts(reportable, "类别")
    system_counts = top_counts(reportable, "系统名称")
    source_counts = top_counts(reportable, "数据来源")
    result_counts = top_counts(reportable, "结果")
    person_counts = top_counts(reportable, "实施人")
    document_summaries, missing_documents = collect_delivery_document_summaries(selected, document_root)
    context = {
        "report_config": report_config,
        "ledger": ledger,
        "document_root": document_root,
        "start_date": start_date,
        "end_date": end_date,
        "selected_count": len(selected),
        "reportable": reportable,
        "completed": completed,
        "continued": continued,
        "risk_records": risk_records,
        "category_counts": category_counts,
        "system_counts": system_counts,
        "source_counts": source_counts,
        "result_counts": result_counts,
        "person_counts": person_counts,
        "document_summaries": document_summaries,
        "missing_documents": missing_documents,
    }

    lines = [
        f"# {report_config['title']}（{start_date.isoformat()} 至 {end_date.isoformat()}）",
        "",
        "## 1. 概览",
        "",
        f"- 台账文件：`{ledger}`",
        f"- 交付文档搜索目录：`{document_root}`",
        f"- 周期内记录数：{len(selected)}",
        f"- 纳入汇报记录数：{len(reportable)}",
        f"- 已完成事项：{len(completed)}",
        f"- 延续事项：{len(continued)}",
        f"- 风险/影响事项：{len(risk_records)}",
        "",
    ]
    lines.extend(render_management_summary(report_config, start_date, end_date, reportable,
                                           completed, continued, risk_records,
                                           category_counts, system_counts, document_summaries))
    lines.extend(render_count_table("2. 按类别统计", category_counts))
    lines.extend(render_count_table("3. 按系统统计", system_counts))
    lines.extend(render_count_table("4. 按数据来源统计", source_counts))
    lines.extend(render_count_table("5. 按结果统计", result_counts))
    lines.extend(render_count_table("6. 按实施人统计", person_counts))
    lines.extend(render_record_table(f"7. {report_config['focus_title']}", reportable, 30))
    lines.extend(render_record_table("8. 风险与问题", risk_records, 20))
    lines.extend(render_record_table("9. 延续事项", continued, 20))
    lines.extend(render_delivery_document_summaries("10. 交付文档重点内容", document_summaries, missing_documents))
    lines.extend(render_delivery_document_list("11. 交付文档清单", document_summaries, missing_documents))
    lines.extend([f"## 12. {report_config['plan_title']}", ""])
    lines.extend(f"- {item}" for item in report_config["plan_items"])
    lines.append("")

    output_dir = TMP_REPORTS
    output_dir.mkdir(parents=True, exist_ok=True)
    file_prefix = report_config["file_prefix"]
    md_path = output_dir / f"{file_prefix}_{start_date.isoformat()}_{end_date.isoformat()}_ai.md"
    word_path = output_dir / f"{file_prefix}_{start_date.isoformat()}_{end_date.isoformat()}_ai.docx"
    base_word_path = output_dir / f"{file_prefix}_{start_date.isoformat()}_{end_date.isoformat()}.docx"

    error_warning = ""
    enhanced_md = "\n".join(lines)
    try:
        prompt = build_report_prompt(context, lines)
        answer = ask_llm(url, model, prompt, api_key=api_key, timeout=300)
        if answer and len(answer) > 50:
            enhanced_md = answer
    except Exception as exc:
        error_warning = f"AI 增强失败（{exc}），已回退为基础版"
        create_docx_report(base_word_path, context)
        md_path.write_text("\n".join(lines), encoding="utf-8")
        return {
            "report_type": report_type,
            "report_title": report_config["title"],
            "report_path": str(md_path),
            "word_path": str(base_word_path),
            "ledger_path": str(ledger),
            "document_root": str(document_root),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "total_records": len(selected),
            "reportable_records": len(reportable),
            "completed_records": len(completed),
            "continued_records": len(continued),
            "risk_records": len(risk_records),
            "category_counts": category_counts,
            "system_counts": system_counts,
            "person_counts": person_counts,
            "delivery_document_count": len(document_summaries),
            "missing_delivery_documents": missing_documents,
            "delivery_document_summaries": document_summaries,
            "preview": "\n".join(lines[:40]),
            "ai_enhanced": False,
            "error_warning": error_warning,
        }

    chart_paths = generate_report_charts(context, output_dir)
    md_path.write_text(enhanced_md, encoding="utf-8")
    create_ai_docx_report(word_path, enhanced_md, context, chart_paths)

    return {
        "report_type": report_type,
        "report_title": report_config["title"],
        "report_path": str(md_path),
        "word_path": str(word_path),
        "ledger_path": str(ledger),
        "document_root": str(document_root),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_records": len(selected),
        "reportable_records": len(reportable),
        "completed_records": len(completed),
        "continued_records": len(continued),
        "risk_records": len(risk_records),
        "category_counts": category_counts,
        "system_counts": system_counts,
        "person_counts": person_counts,
        "delivery_document_count": len(document_summaries),
        "missing_delivery_documents": missing_documents,
        "delivery_document_summaries": document_summaries,
        "preview": enhanced_md[:2000],
        "ai_enhanced": True,
        "chart_count": len(chart_paths),
        "error_warning": error_warning,
    }


def generate_report(report_type, ledger_path, start_date_text, end_date_text, document_root_text=None):
    report_config = REPORT_TYPES.get(report_type)
    if not report_config:
        raise ValueError("报告类型必须是 weekly、monthly、quarterly 或 annual")

    ledger = Path(ledger_path).expanduser().resolve()
    if document_root_text:
        document_root_text = document_root_text.strip().strip("'\"")
    document_root = Path(document_root_text).expanduser().resolve() if document_root_text else ledger.parent
    start_date = parse_iso_date(start_date_text, "开始日期")
    end_date = parse_iso_date(end_date_text, "结束日期")
    if end_date < start_date:
        raise ValueError("结束日期不能早于开始日期")

    records = load_records(ledger)
    selected, reportable = filter_weekly_records(records, start_date, end_date)
    completed = [r for r in reportable if stringify(r.get("结果")) == "完成"]
    continued = [r for r in reportable if stringify(r.get("结果")) == "延续"]
    risk_records = [r for r in reportable
                    if stringify(r.get("业务影响")) in {"已影响", "有风险"}
                    or stringify(r.get("事件级别")) in {"高", "中"}]

    category_counts = top_counts(reportable, "类别")
    system_counts = top_counts(reportable, "系统名称")
    source_counts = top_counts(reportable, "数据来源")
    result_counts = top_counts(reportable, "结果")
    person_counts = top_counts(reportable, "实施人")
    document_summaries, missing_documents = collect_delivery_document_summaries(selected, document_root)
    context = {
        "report_config": report_config,
        "ledger": ledger,
        "document_root": document_root,
        "start_date": start_date,
        "end_date": end_date,
        "selected_count": len(selected),
        "reportable": reportable,
        "completed": completed,
        "continued": continued,
        "risk_records": risk_records,
        "category_counts": category_counts,
        "system_counts": system_counts,
        "source_counts": source_counts,
        "result_counts": result_counts,
        "person_counts": person_counts,
        "document_summaries": document_summaries,
        "missing_documents": missing_documents,
    }

    lines = [
        f"# {report_config['title']}（{start_date.isoformat()} 至 {end_date.isoformat()}）",
        "",
        "## 1. 概览",
        "",
        f"- 台账文件：`{ledger}`",
        f"- 交付文档搜索目录：`{document_root}`",
        f"- 周期内记录数：{len(selected)}",
        f"- 纳入汇报记录数：{len(reportable)}",
        f"- 已完成事项：{len(completed)}",
        f"- 延续事项：{len(continued)}",
        f"- 风险/影响事项：{len(risk_records)}",
        "",
    ]
    lines.extend(render_management_summary(report_config, start_date, end_date, reportable,
                                           completed, continued, risk_records,
                                           category_counts, system_counts, document_summaries))
    lines.extend(render_count_table("2. 按类别统计", category_counts))
    lines.extend(render_count_table("3. 按系统统计", system_counts))
    lines.extend(render_count_table("4. 按数据来源统计", source_counts))
    lines.extend(render_count_table("5. 按结果统计", result_counts))
    lines.extend(render_count_table("6. 按实施人统计", person_counts))
    lines.extend(render_record_table(f"7. {report_config['focus_title']}", reportable, 30))
    lines.extend(render_record_table("8. 风险与问题", risk_records, 20))
    lines.extend(render_record_table("9. 延续事项", continued, 20))
    lines.extend(render_delivery_document_summaries("10. 交付文档重点内容", document_summaries, missing_documents))
    lines.extend(render_delivery_document_list("11. 交付文档清单", document_summaries, missing_documents))
    lines.extend([f"## 12. {report_config['plan_title']}", ""])
    lines.extend(f"- {item}" for item in report_config["plan_items"])
    lines.append("")

    output_dir = TMP_REPORTS
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{report_config['file_prefix']}_{start_date.isoformat()}_{end_date.isoformat()}.md"
    word_path = output_dir / f"{report_config['file_prefix']}_{start_date.isoformat()}_{end_date.isoformat()}.docx"
    output_path.write_text("\n".join(lines), encoding="utf-8")
    create_docx_report(word_path, context)

    return {
        "report_type": report_type,
        "report_title": report_config["title"],
        "report_path": str(output_path),
        "word_path": str(word_path),
        "ledger_path": str(ledger),
        "document_root": str(document_root),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_records": len(selected),
        "reportable_records": len(reportable),
        "completed_records": len(completed),
        "continued_records": len(continued),
        "risk_records": len(risk_records),
        "category_counts": category_counts,
        "system_counts": system_counts,
        "person_counts": person_counts,
        "delivery_document_count": len(document_summaries),
        "missing_delivery_documents": missing_documents,
        "delivery_document_summaries": document_summaries,
        "preview": "\n".join(lines[:40]),
    }


def generate_weekly_report(ledger_path, start_date_text, end_date_text, document_root_text=None):
    return generate_report("weekly", ledger_path, start_date_text, end_date_text, document_root_text)


def generate_monthly_report(ledger_path, start_date_text, end_date_text, document_root_text=None):
    return generate_report("monthly", ledger_path, start_date_text, end_date_text, document_root_text)


def generate_quarterly_report(ledger_path, start_date_text, end_date_text, document_root_text=None):
    return generate_report("quarterly", ledger_path, start_date_text, end_date_text, document_root_text)


def generate_annual_report(ledger_path, start_date_text, end_date_text, document_root_text=None):
    return generate_report("annual", ledger_path, start_date_text, end_date_text, document_root_text)
